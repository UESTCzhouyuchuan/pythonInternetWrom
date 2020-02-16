# -*- coding:utf-8 -*-
# 文件操作
'''
包括：
writeToTxt(filename, arr):写入到TXT文件中
writeToXlsx(filename, date, title_EN, title_ZH, type, row):写入到excel表格中
writeToFile(filename,data):把data写入到文件中
getFileData(filename):从文件中获得数据
'''
import openpyxl
from printState import *
import json


def writeToTxt(filename, arr):
    file = open(filename, mode='w',encoding="utf8")
    # str的join方法
    _str = "\n\n".join(arr)
    try:
        file.write(_str)
    except Exception as e:
        printError('写入文件' + filename+e)
    else:
        printSuccess('写入文件' + filename)
    file.flush()
    file.close()

def writeToXlsx(filename, date, title_EN, title_ZH, type, row):
    try:
        sorFile = openpyxl.load_workbook(filename)
    except Exception as e:
        printError("打开XlSX" + filename + "失败" + e)
    ws = sorFile[sorFile.sheetnames[0]]
    ws.cell(row, 1).value = int(date)
    ws.cell(row, 2).value = title_EN
    ws.cell(row, 3).value = title_ZH
    ws.cell(row, 4).value = type
    try:
        sorFile.save(filename)
    except Exception as e:
        printError("写入第" + str(row) + "行到XLSX" + filename + e)
    else:
        printSuccess("写入第" + str(row) + "行到XLSX" + filename)
    sorFile.close()
def writeToFile(filename,data):
    try:
        file = open(filename, mode='w+', encoding='UTF-8')
    except Exception:
        print("Error can't open")
    else:
        file.write(str(json.dumps(data, ensure_ascii=False, indent=2)))
        file.close()
        printSuccess('数据写入文件'+filename)
def getFileData(filename):
    ret = None
    try:
        with open(filename, 'r', encoding="utf-8") as f:
            ret =  json.load(f)
            f.close()
    except Exception as e:
        printError("打开文件"+filename+str(e))
    else:
        return ret
def main():
    ret = getFileData("test.json")
if __name__ == "__main__":
    # filename = "./corpus on belt and road initiative/corpus on belt and road initiative/Title/Title_Single.xlsx"
    # writeToXlsx(filename, "20200214", "EN", "汉语", "China", 2);
    main()