import json
import re
from multiprocessing import Pool

import openpyxl

# 进程池
def pool(data, cb, processNumn=10):
    try:
        printTip(f"开始多进程,一共{len(data)}个数据, {processNumn}个进程")
        printTip("开始爬取....")
        p = Pool(processNumn)
        ret = p.starmap(cb, data)
        return ret
    except Exception as e:
        printError(f"开启多进程, error: {e}")
# 常见函数
def make_json(s):
    splits = s.strip().split('\n')
    ret = {}
    for item in splits:
        t = item.split(': ')
        ret[t[0]] = t[1]
    return ret
def re_strip(s, t=r'\s'):
    t_format = rf'^{t}*|{t}*$'
    s_re = re.compile(t_format)
    s = s_re.sub('', s)
    return s


def ILLEGAL_CHARACTERS(text):
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    text = ILLEGAL_CHARACTERS_RE.sub(r'', text)
    text = text.replace('<![CDATA[<div>', '')
    text = text.replace('\xa0', ' ')
    text = text.strip()
    return text

# 文件操作

def writeToFile(filename, data, encoding="UTF-8"):
    try:
        file = open(filename, mode='w+', encoding=encoding)
    except Exception as e:
        printError(f"can't open {filename}, error: {e}")
    else:
        json.dump(data, file, ensure_ascii=False, indent=2)
        file.close()
        printSuccess('数据写入文件' + filename)


def readFromFile(filename, encoding="UTF-8"):
    try:
        file = open(filename, mode='r', encoding=encoding)
    except Exception as e:
        printError(f"can't open {filename}, error: {e}")
    else:
        data = json.load(file)
        file.close()
        printSuccess('读入数据文件' + filename)
        return data





def writeToXlsx(filename, data, rows = None):
    try:
        wb = openpyxl.Workbook()
    except Exception as e:
        printError(f"创建{filename},error: {e}")
    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    if rows is None:
        headers = list(data[0].keys())
    else:
        headers = rows[:]
    headers.insert(0, 'number')
    for i in range(len(headers)):
        ws.cell(1, i + 1).value = headers[i]
        ws.cell(1, i + 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    row = 0
    for i in range(len(data)):
        item = data[i]
        if item is not None:
            try:
                ws.cell(row + 2, 1).value = row + 1
                ws.cell(row + 2, 1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

                for j in range(1, len(headers)):
                    ws.cell(row + 2, j + 1).value = str(item[headers[j]])
                row += 1
            except Exception as e:
                printError(f"写入第{row}行数据到{filename},{item}, error: {e}")
    try:
        wb.save(filename)
    except Exception as e:
        printError(f"保存文件{filename}, error: {e}")
    else:
        printSuccess(f"写入数据到{filename}")
    wb.close()

# 封装输出格式化
'''
printError(_str):输出错误信息
printSuccess(_str):输出执行成功的提示信息
printInfo(_str):输出普通信息
printTip(_str):输出提示信息
'''


def printError(_str):
    print("\033[1;31mError：{}出现错误\033[0m".format(_str))


def printSuccess(_str):
    print("\033[1;32mSuccess：{}成功\033[0m".format(_str))


def printInfo(_str):
    print("\033[1;34mInfo：{}\033[0m".format(_str))


def printTip(_str):
    print("\033[0;36mIip：{}\033[0m".format(_str))

